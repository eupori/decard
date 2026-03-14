import asyncio
import csv
import io
import logging
import re
import uuid
from datetime import timedelta
from typing import Optional

from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .auth import get_device_id, get_owner_filter, get_owner_filter_for_folder, get_owner_id
from .claude_cli import release_session_semaphore
from .config import settings
from .database import get_db, SessionLocal
from .models import (
    SessionModel, CardModel, GradeModel, FolderModel, CardReviewModel,
    PublicCardsetModel, PublicCardModel,
    CardResponse, CardUpdate, SessionResponse, GradeResponse,
    FolderCreate, FolderUpdate, FolderResponse, SaveToLibraryRequest,
    ManualCardInput, ManualSessionCreate,
    ReviewRequest, ReviewResponse, StudyStatsResponse,
    PublishRequest,
)
from .pdf_service import extract_text_from_pdf, validate_pdf
from .billing_service import get_billing_status, can_generate
from .card_service import generate_cards
from .grade_service import grade_answer
from .srs_service import calculate_sm2

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/v1")


# ──────────────────────────────────────
# POST /api/v1/generate — PDF 업로드 + 카드 생성
# ──────────────────────────────────────

@router.post("/generate")
async def generate(
    request: Request,
    file: UploadFile = File(...),
    template_type: str = Form("definition"),
    db: Session = Depends(get_db),
    device_id: str = Depends(get_device_id),
):
    import time
    t0 = time.time()

    # "subjective" — TODO: MVP 이후 추가
    if template_type not in ("definition", "cloze", "comparison"):
        raise HTTPException(400, "지원하지 않는 템플릿입니다. (definition / cloze / comparison)")

    # 월간 생성 한도 체크
    owner = get_owner_id(request)
    billing = can_generate(owner["user_id"], device_id, db)
    if not billing["allowed"]:
        raise HTTPException(429, billing["message"])

    # File size pre-check (Content-Length header)
    if file.size and file.size > settings.MAX_PDF_SIZE_MB * 1024 * 1024:
        raise HTTPException(400, f"파일 크기가 {settings.MAX_PDF_SIZE_MB}MB를 초과합니다.")

    t1 = time.time()
    content = await file.read()
    t2 = time.time()

    # PDF 검증 (첫 페이지만 빠르게)
    valid, error = validate_pdf(content, settings.MAX_PDF_SIZE_MB)
    if not valid:
        raise HTTPException(400, error)

    t3 = time.time()

    # 동시 처리 세션 수 제한
    processing_count = db.query(SessionModel).filter(
        SessionModel.status == "processing"
    ).count()
    if processing_count >= settings.MAX_CONCURRENT_SESSIONS:
        raise HTTPException(
            429,
            detail={
                "error": "서버가 바쁩니다. 잠시 후 다시 시도해주세요.",
                "processing_count": processing_count,
                "max_concurrent": settings.MAX_CONCURRENT_SESSIONS,
                "retry_after_seconds": 30,
            },
        )

    # 세션 생성 (즉시 반환 — 텍스트 추출은 백그라운드에서)
    session = SessionModel(
        filename=re.sub(r'<[^>]+>', '', file.filename or "unknown.pdf"),
        page_count=0,
        template_type=template_type,
        device_id=device_id,
        user_id=owner["user_id"],
        status="processing",
    )
    db.add(session)
    db.commit()
    db.refresh(session)

    t4 = time.time()

    # 백그라운드에서 텍스트 추출 + 카드 생성
    asyncio.create_task(
        _generate_in_background(session.id, content, template_type)
    )

    logger.info(
        "POST /generate 타이밍: file.read=%.2fs, validate=%.2fs, db=%.2fs, total=%.2fs, size=%.1fMB",
        t2 - t1, t3 - t2, t4 - t3, t4 - t0, len(content) / 1024 / 1024,
    )

    return _build_session_response(session)


# ──────────────────────────────────────
# GET /api/v1/billing/status — 월간 사용량 조회
# ──────────────────────────────────────

@router.get("/billing/status")
def billing_status(
    request: Request,
    db: Session = Depends(get_db),
    device_id: str = Depends(get_device_id),
):
    owner = get_owner_id(request)
    return get_billing_status(owner["user_id"], device_id, db)


# ──────────────────────────────────────
# GET /api/v1/sessions — 세션 목록
# ──────────────────────────────────────

@router.get("/sessions")
def list_sessions(request: Request, db: Session = Depends(get_db)):
    owner_filter = get_owner_filter(request)
    query = db.query(SessionModel).filter(
        SessionModel.status.in_(["completed", "processing", "failed"])
    )
    sessions = (
        owner_filter(query)
        .order_by(SessionModel.created_at.desc())
        .limit(50)
        .all()
    )
    result = []
    for s in sessions:
        item = {
            "id": s.id,
            "filename": s.filename,
            "page_count": s.page_count,
            "template_type": s.template_type,
            "status": s.status,
            "card_count": len(s.cards),
            "folder_id": s.folder_id,
            "display_name": s.display_name,
            "source_type": s.source_type or "pdf",
            "progress": s.progress or 0,
            "total_chunks": s.total_chunks or 0,
            "completed_chunks": s.completed_chunks or 0,
            "created_at": s.created_at.isoformat() + "Z",
        }
        if s.error_message:
            item["error_message"] = s.error_message
        result.append(item)
    return result


# ──────────────────────────────────────
# DELETE /api/v1/sessions/{id} — 세션 삭제
# ──────────────────────────────────────

@router.delete("/sessions/{session_id}")
def delete_session(session_id: str, request: Request, db: Session = Depends(get_db)):
    owner_filter = get_owner_filter(request)
    session = owner_filter(db.query(SessionModel).filter(SessionModel.id == session_id)).first()
    if not session:
        raise HTTPException(404, "세션을 찾을 수 없습니다.")
    db.delete(session)
    db.commit()
    return {"deleted": session_id}


# ──────────────────────────────────────
# GET /api/v1/sessions/{id} — 세션 조회
# ──────────────────────────────────────

@router.get("/sessions/{session_id}")
def get_session(session_id: str, request: Request, db: Session = Depends(get_db)):
    owner_filter = get_owner_filter(request)
    session = owner_filter(db.query(SessionModel).filter(SessionModel.id == session_id)).first()
    if not session:
        raise HTTPException(404, "세션을 찾을 수 없습니다.")
    return _build_session_response(session)


# ──────────────────────────────────────
# PATCH /api/v1/cards/{id} — 카드 상태/내용 수정
# ──────────────────────────────────────

@router.patch("/cards/{card_id}")
def update_card(card_id: str, update: CardUpdate, db: Session = Depends(get_db)):
    card = db.query(CardModel).filter(CardModel.id == card_id).first()
    if not card:
        raise HTTPException(404, "카드를 찾을 수 없습니다.")

    if update.status and update.status in ("accepted", "rejected", "pending"):
        card.status = update.status
    if update.front is not None:
        card.front = update.front
    if update.back is not None:
        card.back = update.back

    db.commit()
    db.refresh(card)

    return _card_to_response(card)


# ──────────────────────────────────────
# POST /api/v1/sessions/{id}/accept-all — 전체 채택
# ──────────────────────────────────────

@router.post("/sessions/{session_id}/accept-all")
def accept_all(session_id: str, db: Session = Depends(get_db)):
    cards = db.query(CardModel).filter(
        CardModel.session_id == session_id,
        CardModel.status == "pending",
    ).all()
    for card in cards:
        card.status = "accepted"
    db.commit()
    return {"accepted": len(cards)}


# ──────────────────────────────────────
# POST /api/v1/cards/{id}/grade — AI 채점
# ──────────────────────────────────────

@router.post("/cards/{card_id}/grade")
async def grade_card(
    card_id: str,
    user_answer: str = Form(""),
    drawing: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
):
    card = db.query(CardModel).filter(CardModel.id == card_id).first()
    if not card:
        raise HTTPException(404, "카드를 찾을 수 없습니다.")

    # 손글씨 이미지 읽기
    drawing_image = None
    has_drawing = False
    if drawing and drawing.filename:
        drawing_image = await drawing.read()
        has_drawing = True

    if not user_answer.strip() and not drawing_image:
        raise HTTPException(400, "답안을 입력해주세요. (텍스트 또는 손글씨)")

    try:
        result = await grade_answer(
            question=card.front,
            model_answer=card.back,
            user_answer=user_answer,
            drawing_image=drawing_image,
        )

        grade = GradeModel(
            card_id=card_id,
            user_answer=user_answer,
            has_drawing=has_drawing,
            score=result["score"],
            feedback=result["feedback"],
        )
        db.add(grade)
        db.commit()
        db.refresh(grade)

        return GradeResponse(
            id=grade.id,
            card_id=card_id,
            user_answer=user_answer,
            score=result["score"],
            feedback=result["feedback"],
            model_answer=card.back,
        ).model_dump()

    except Exception as e:
        logger.exception("채점 실패: card=%s", card_id)
        raise HTTPException(500, _safe_error("채점에 실패했습니다", e))


# ──────────────────────────────────────
# GET /api/v1/sessions/{id}/download — CSV 다운로드
# ──────────────────────────────────────

@router.get("/sessions/{session_id}/download")
def download_csv(session_id: str, request: Request, db: Session = Depends(get_db)):
    owner_filter = get_owner_filter(request)
    session = owner_filter(db.query(SessionModel).filter(SessionModel.id == session_id)).first()
    if not session:
        raise HTTPException(404, "세션을 찾을 수 없습니다.")

    # 채택 카드 우선, 없으면 pending 포함
    cards = db.query(CardModel).filter(
        CardModel.session_id == session_id,
        CardModel.status.in_(["accepted", "pending"]),
    ).all()

    if not cards:
        raise HTTPException(404, "다운로드할 카드가 없습니다.")

    # Anki 임포트 포맷 (TSV: 앞면 \t 뒷면 \t 태그)
    output = io.StringIO()
    writer = csv.writer(output, delimiter="\t")

    for card in cards:
        back_with_evidence = (
            f"{card.back}\n\n"
            f"📖 근거 (p.{card.evidence_page}): {card.evidence}"
        )
        writer.writerow([card.front, back_with_evidence, card.tags])

    output.seek(0)
    safe_name = session.filename.replace(".pdf", "").replace(" ", "_")
    filename = f"decard_{safe_name}_{session.template_type}.txt"

    # RFC 5987: 한국어 파일명을 UTF-8로 인코딩
    from urllib.parse import quote
    encoded_filename = quote(filename)

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/tab-separated-values",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )


# ──────────────────────────────────────
# Folder CRUD
# ──────────────────────────────────────

PRESET_COLORS = {"#C2E7DA", "#6290C3", "#9B72CF", "#F59E0B", "#EF4444", "#94A3B8"}


@router.get("/folders")
def list_folders(request: Request, db: Session = Depends(get_db)):
    folder_filter = get_owner_filter_for_folder(request)
    folders = (
        folder_filter(db.query(FolderModel))
        .order_by(FolderModel.created_at.desc())
        .all()
    )
    result = []
    for f in folders:
        session_count = len(f.sessions)
        card_count = sum(len(s.cards) for s in f.sessions)
        result.append({
            "id": f.id,
            "name": f.name,
            "color": f.color,
            "session_count": session_count,
            "card_count": card_count,
            "created_at": f.created_at.isoformat() + "Z",
            "updated_at": f.updated_at.isoformat() + "Z",
        })
    return result


@router.post("/folders")
def create_folder(
    body: FolderCreate,
    request: Request,
    db: Session = Depends(get_db),
    device_id: str = Depends(get_device_id),
):
    owner = get_owner_id(request)
    color = body.color if body.color and body.color in PRESET_COLORS else "#C2E7DA"
    folder = FolderModel(
        name=body.name,
        color=color,
        user_id=owner["user_id"],
        device_id=device_id,
    )
    db.add(folder)
    db.commit()
    db.refresh(folder)
    return {
        "id": folder.id,
        "name": folder.name,
        "color": folder.color,
        "session_count": 0,
        "card_count": 0,
        "created_at": folder.created_at.isoformat() + "Z",
        "updated_at": folder.updated_at.isoformat() + "Z",
    }


@router.patch("/folders/{folder_id}")
def update_folder(
    folder_id: str,
    body: FolderUpdate,
    request: Request,
    db: Session = Depends(get_db),
):
    folder_filter = get_owner_filter_for_folder(request)
    folder = folder_filter(db.query(FolderModel).filter(FolderModel.id == folder_id)).first()
    if not folder:
        raise HTTPException(404, "폴더를 찾을 수 없습니다.")

    if body.name is not None:
        folder.name = body.name
    if body.color is not None and body.color in PRESET_COLORS:
        folder.color = body.color
    from datetime import datetime as dt
    folder.updated_at = dt.utcnow()

    db.commit()
    db.refresh(folder)

    session_count = len(folder.sessions)
    card_count = sum(len(s.cards) for s in folder.sessions)
    return {
        "id": folder.id,
        "name": folder.name,
        "color": folder.color,
        "session_count": session_count,
        "card_count": card_count,
        "created_at": folder.created_at.isoformat() + "Z",
        "updated_at": folder.updated_at.isoformat() + "Z",
    }


@router.delete("/folders/{folder_id}")
def delete_folder(
    folder_id: str,
    request: Request,
    db: Session = Depends(get_db),
):
    folder_filter = get_owner_filter_for_folder(request)
    folder = folder_filter(db.query(FolderModel).filter(FolderModel.id == folder_id)).first()
    if not folder:
        raise HTTPException(404, "폴더를 찾을 수 없습니다.")

    # 세션의 folder_id를 null로 (세션 자체는 보존)
    for s in folder.sessions:
        s.folder_id = None
    db.delete(folder)
    db.commit()
    return {"deleted": folder_id}


@router.get("/folders/{folder_id}/sessions")
def list_folder_sessions(
    folder_id: str,
    request: Request,
    db: Session = Depends(get_db),
):
    folder_filter = get_owner_filter_for_folder(request)
    folder = folder_filter(db.query(FolderModel).filter(FolderModel.id == folder_id)).first()
    if not folder:
        raise HTTPException(404, "폴더를 찾을 수 없습니다.")

    sessions = (
        db.query(SessionModel)
        .filter(SessionModel.folder_id == folder_id)
        .order_by(SessionModel.created_at.desc())
        .all()
    )
    result = []
    for s in sessions:
        item = {
            "id": s.id,
            "filename": s.filename,
            "page_count": s.page_count,
            "template_type": s.template_type,
            "status": s.status,
            "card_count": len(s.cards),
            "folder_id": s.folder_id,
            "display_name": s.display_name,
            "source_type": s.source_type or "pdf",
            "progress": s.progress or 0,
            "total_chunks": s.total_chunks or 0,
            "completed_chunks": s.completed_chunks or 0,
            "created_at": s.created_at.isoformat() + "Z",
        }
        if s.error_message:
            item["error_message"] = s.error_message
        result.append(item)
    return result


# ──────────────────────────────────────
# Save to / Remove from Library
# ──────────────────────────────────────

@router.post("/sessions/{session_id}/save-to-library")
def save_to_library(
    session_id: str,
    body: SaveToLibraryRequest,
    request: Request,
    db: Session = Depends(get_db),
    device_id: str = Depends(get_device_id),
):
    owner_filter = get_owner_filter(request)
    session = owner_filter(db.query(SessionModel).filter(SessionModel.id == session_id)).first()
    if not session:
        raise HTTPException(404, "세션을 찾을 수 없습니다.")

    # 폴더 결정: 기존 폴더 or 새 폴더 생성
    if body.folder_id:
        folder_filter = get_owner_filter_for_folder(request)
        folder = folder_filter(db.query(FolderModel).filter(FolderModel.id == body.folder_id)).first()
        if not folder:
            raise HTTPException(404, "폴더를 찾을 수 없습니다.")
    elif body.new_folder_name:
        owner = get_owner_id(request)
        color = body.new_folder_color if body.new_folder_color and body.new_folder_color in PRESET_COLORS else "#C2E7DA"
        folder = FolderModel(
            name=body.new_folder_name,
            color=color,
            user_id=owner["user_id"],
            device_id=device_id,
        )
        db.add(folder)
        db.flush()
    else:
        raise HTTPException(400, "folder_id 또는 new_folder_name을 지정해주세요.")

    session.folder_id = folder.id
    if body.display_name is not None:
        session.display_name = body.display_name

    db.commit()
    return {
        "session_id": session.id,
        "folder_id": folder.id,
        "folder_name": folder.name,
        "display_name": session.display_name,
    }


@router.delete("/sessions/{session_id}/remove-from-library")
def remove_from_library(
    session_id: str,
    request: Request,
    db: Session = Depends(get_db),
):
    owner_filter = get_owner_filter(request)
    session = owner_filter(db.query(SessionModel).filter(SessionModel.id == session_id)).first()
    if not session:
        raise HTTPException(404, "세션을 찾을 수 없습니다.")

    session.folder_id = None
    session.display_name = None
    db.commit()
    return {"removed": session_id}


# ──────────────────────────────────────
# POST /api/v1/sessions/{id}/share — 공유 링크 생성
# ──────────────────────────────────────

@router.post("/sessions/{session_id}/share")
def create_share_link(
    session_id: str,
    request: Request,
    db: Session = Depends(get_db),
):
    owner_filter = get_owner_filter(request)
    session = owner_filter(db.query(SessionModel).filter(SessionModel.id == session_id)).first()
    if not session:
        raise HTTPException(404, "세션을 찾을 수 없습니다.")
    if session.status != "completed":
        raise HTTPException(400, "완료된 세션만 공유할 수 있습니다.")

    # 이미 share_key가 있으면 그대로 반환
    if not session.share_key:
        session.share_key = uuid.uuid4().hex[:12]
        db.commit()
        db.refresh(session)

    return {
        "share_key": session.share_key,
        "share_url": f"{settings.FRONTEND_URL}/#/shared/{session.share_key}",
    }


# ──────────────────────────────────────
# GET /api/v1/shared/{share_key} — 공유 세션 읽기 전용 조회
# ──────────────────────────────────────

@router.get("/shared/{share_key}")
def get_shared_session(share_key: str, db: Session = Depends(get_db)):
    session = db.query(SessionModel).filter(SessionModel.share_key == share_key).first()
    if not session:
        raise HTTPException(404, "공유 링크를 찾을 수 없습니다.")

    # accepted 카드만 공개 (pending/rejected는 비공개)
    cards = [
        {
            "id": c.id,
            "front": c.front,
            "back": c.back,
            "evidence": c.evidence,
            "evidence_page": c.evidence_page,
            "tags": c.tags,
            "template_type": c.template_type,
        }
        for c in session.cards if c.status == "accepted"
    ]
    return {
        "title": session.display_name or session.filename.replace(".pdf", ""),
        "card_count": len(cards),
        "template_type": session.template_type,
        "created_at": session.created_at.isoformat() + "Z",
        "cards": cards,
        "is_shared": True,  # 프론트에서 읽기 전용 모드 판별용
    }


# ──────────────────────────────────────
# POST /api/v1/sessions/create-manual — 수동 카드 만들기
# ──────────────────────────────────────

@router.post("/sessions/create-manual")
def create_manual_session(
    body: ManualSessionCreate,
    request: Request,
    db: Session = Depends(get_db),
    device_id: str = Depends(get_device_id),
):
    if not body.cards or len(body.cards) < 1:
        raise HTTPException(400, "카드를 최소 1장 이상 입력해주세요.")
    if len(body.cards) > 200:
        raise HTTPException(400, "카드는 최대 200장까지 입력할 수 있습니다.")

    ALLOWED_MANUAL_TYPES = {"definition", "multiple_choice", "cloze"}
    for card_input in body.cards:
        if card_input.template_type not in ALLOWED_MANUAL_TYPES:
            raise HTTPException(400, f"지원하지 않는 카드 유형입니다: {card_input.template_type}")

    # 세션 template_type: 첫 번째 카드 유형 사용 (혼합 가능)
    session_template = body.cards[0].template_type if body.cards else "definition"

    owner = get_owner_id(request)
    session = SessionModel(
        filename=body.display_name or "직접 입력",
        page_count=0,
        template_type=session_template,
        device_id=device_id,
        user_id=owner["user_id"],
        source_type="manual",
        status="completed",
    )
    db.add(session)
    db.flush()

    for card_input in body.cards:
        card = CardModel(
            session_id=session.id,
            front=card_input.front,
            back=card_input.back,
            evidence=card_input.evidence or "",
            evidence_page=0,
            template_type=card_input.template_type,
            status="accepted",
        )
        db.add(card)

    db.commit()
    db.refresh(session)
    return _build_session_response(session)


# ──────────────────────────────────────
# POST /api/v1/sessions/import-file — CSV/XLSX 파일 임포트
# ──────────────────────────────────────

@router.post("/sessions/import-file")
async def import_file(
    request: Request,
    file: UploadFile = File(...),
    display_name: str = Form(None),
    db: Session = Depends(get_db),
    device_id: str = Depends(get_device_id),
):
    # 파일 크기 제한 (5MB)
    max_size = 5 * 1024 * 1024
    if file.size and file.size > max_size:
        raise HTTPException(400, "파일 크기가 5MB를 초과합니다.")

    content = await file.read()
    if len(content) > max_size:
        raise HTTPException(400, "파일 크기가 5MB를 초과합니다.")

    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        parsed = _parse_csv(content)
        source_type = "csv"
    elif ext == "xlsx":
        parsed = _parse_xlsx(content)
        source_type = "xlsx"
    else:
        raise HTTPException(400, "지원하지 않는 파일 형식입니다. (csv, xlsx)")

    if not parsed:
        raise HTTPException(400, "파일에서 카드를 추출할 수 없습니다.")
    if len(parsed) > 500:
        raise HTTPException(400, "카드는 최대 500장까지 가져올 수 있습니다.")

    owner = get_owner_id(request)
    session = SessionModel(
        filename=display_name or filename or "파일 임포트",
        page_count=0,
        template_type="definition",
        device_id=device_id,
        user_id=owner["user_id"],
        source_type=source_type,
        status="completed",
    )
    db.add(session)
    db.flush()

    for row in parsed:
        card = CardModel(
            session_id=session.id,
            front=row["front"],
            back=row["back"],
            evidence=row.get("evidence", ""),
            evidence_page=0,
            template_type="definition",
            status="accepted",
        )
        db.add(card)

    db.commit()
    db.refresh(session)
    return _build_session_response(session)


_HEADER_FRONT = {"front", "앞면", "질문", "question"}
_HEADER_BACK = {"back", "뒷면", "답", "답변", "answer"}


def _parse_csv(content: bytes) -> list[dict]:
    """CSV 파싱 — BOM 처리, 헤더 자동 감지."""
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []

    # 헤더 감지
    first_row = [cell.strip().lower() for cell in rows[0]]
    has_header = bool(set(first_row) & _HEADER_FRONT) or bool(set(first_row) & _HEADER_BACK)

    data_rows = rows[1:] if has_header else rows
    result = []
    for row in data_rows:
        if len(row) < 2:
            continue
        front = row[0].strip()
        back = row[1].strip()
        if not front or not back:
            continue
        evidence = row[2].strip() if len(row) > 2 else ""
        result.append({"front": front, "back": back, "evidence": evidence})
    return result


def _parse_xlsx(content: bytes) -> list[dict]:
    """XLSX 파싱 — openpyxl 사용, 헤더 자동 감지."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    # 헤더 감지
    first_row = [str(cell or "").strip().lower() for cell in rows[0]]
    has_header = bool(set(first_row) & _HEADER_FRONT) or bool(set(first_row) & _HEADER_BACK)

    data_rows = rows[1:] if has_header else rows
    result = []
    for row in data_rows:
        if len(row) < 2:
            continue
        front = str(row[0] or "").strip()
        back = str(row[1] or "").strip()
        if not front or not back:
            continue
        evidence = str(row[2] or "").strip() if len(row) > 2 and row[2] else ""
        result.append({"front": front, "back": back, "evidence": evidence})
    return result


# ──────────────────────────────────────
# Helpers
# ──────────────────────────────────────

async def _generate_in_background(
    session_id: str,
    pdf_content: bytes,
    template_type: str,
) -> None:
    """Request 스코프 밖에서 별도 DB 세션으로 텍스트 추출 + 카드 생성."""
    db = SessionLocal()

    async def _update_progress(completed_chunks: int, total_chunks: int, phase: str):
        """청크 완료 시마다 DB에 진행률 업데이트."""
        try:
            session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
            if not session:
                logger.warning("진행률 업데이트: 세션 없음 session=%s (이미 삭제?)", session_id)
                return
            session.completed_chunks = completed_chunks
            session.total_chunks = total_chunks
            if phase == "extracting":
                session.progress = 5
            elif phase == "chunked":
                session.progress = 15
            elif phase == "generating":
                # 15~85% 구간: 청크 진행에 비례
                session.progress = 15 + int(completed_chunks / max(total_chunks, 1) * 70)
            elif phase == "reviewing":
                session.progress = 85
            elif phase == "done":
                session.progress = 100
            db.commit()
            logger.info("진행률 업데이트: session=%s, phase=%s, %d/%d, progress=%d%%",
                        session_id, phase, completed_chunks, total_chunks, session.progress)
        except Exception as e:
            logger.error("진행률 업데이트 실패: session=%s, error=%s: %s", session_id, type(e).__name__, e)

    try:
        # 텍스트 추출 (백그라운드에서 실행)
        await _update_progress(0, 0, "extracting")
        extraction = extract_text_from_pdf(pdf_content)
        pages = extraction["pages"]
        extraction_method = extraction["method"]
        logger.info("텍스트 추출 완료: method=%s, is_math=%s, pages=%d",
                     extraction_method, extraction.get("is_math"), len(pages))
        if not pages:
            raise ValueError("텍스트를 추출할 수 없는 PDF입니다.")
        if len(pages) > settings.MAX_PAGES:
            raise ValueError(f"페이지 수 초과: {len(pages)}/{settings.MAX_PAGES}")

        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            logger.error("백그라운드 생성: 세션 없음 session=%s", session_id)
            return

        session.page_count = len(pages)
        db.commit()
        await _update_progress(0, 0, "chunked")

        cards_data = await generate_cards(
            pages, template_type,
            session_id=session_id,
            on_progress=_update_progress,
            is_math=extraction.get("is_math", False),
        )

        for card_data in cards_data:
            status = card_data.pop("status", "pending")
            card_data.pop("recommend", None)
            db.add(CardModel(session_id=session.id, status=status, **card_data))

        session.status = "completed"
        session.progress = 100
        db.commit()
        logger.info("백그라운드 생성 완료: session=%s, cards=%d", session_id, len(cards_data))
    except Exception as e:
        logger.exception("백그라운드 카드 생성 실패: session=%s, error=%s: %s", session_id, type(e).__name__, e)
        try:
            session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
            if session:
                session.status = "failed"
                # 사용자에게 보여줄 에러 메시지 — 기술적 세부사항은 줄이고 사유 위주
                err_str = str(e)
                if "타임아웃" in err_str or "timeout" in err_str.lower():
                    session.error_message = "AI 처리 시간이 초과되었습니다. 더 짧은 PDF로 시도해주세요."
                elif "빈 응답" in err_str:
                    session.error_message = "AI가 응답하지 않았습니다. 잠시 후 다시 시도해주세요."
                elif "텍스트를 추출" in err_str:
                    session.error_message = "PDF에서 텍스트를 읽을 수 없습니다. 스캔된 PDF는 지원하지 않습니다."
                elif "페이지 수 초과" in err_str:
                    session.error_message = err_str
                else:
                    session.error_message = f"카드 생성 중 오류가 발생했습니다: {err_str[:200]}"
                db.commit()
                logger.info("세션 실패 저장 완료: session=%s, error_message=%s", session_id, session.error_message)
        except Exception as db_err:
            logger.error("세션 실패 상태 업데이트 불가: session=%s, db_error=%s: %s", session_id, type(db_err).__name__, db_err)
        from .slack import send_slack_alert
        await send_slack_alert(
            "카드 생성 실패",
            f"session: `{session_id}`\nerror: {type(e).__name__}: {e}",
        )
    finally:
        release_session_semaphore(session_id)
        db.close()


# ──────────────────────────────────────
# SRS — 간격 반복 학습
# ──────────────────────────────────────

@router.post("/cards/{card_id}/review")
def review_card(
    card_id: str,
    body: ReviewRequest,
    request: Request,
    db: Session = Depends(get_db),
    device_id: str = Depends(get_device_id),
):
    """카드 복습 결과 기록 (SM-2)."""
    card = db.query(CardModel).filter(CardModel.id == card_id).first()
    if not card:
        raise HTTPException(404, "카드를 찾을 수 없습니다.")

    if body.rating < 1 or body.rating > 4:
        raise HTTPException(400, "rating은 1~4 사이여야 합니다.")

    owner = get_owner_id(request)

    # 이전 복습 기록 조회 (최신 1건)
    prev_review = (
        db.query(CardReviewModel)
        .filter(CardReviewModel.card_id == card_id)
        .order_by(CardReviewModel.reviewed_at.desc())
        .first()
    )

    prev_interval = prev_review.interval_days if prev_review else 0
    prev_ease = prev_review.ease_factor if prev_review else 2.5

    new_interval, new_ease, due_date = calculate_sm2(
        body.rating, prev_interval, prev_ease
    )

    review = CardReviewModel(
        card_id=card_id,
        user_id=owner["user_id"],
        device_id=device_id,
        rating=body.rating,
        interval_days=new_interval,
        ease_factor=new_ease,
        due_date=due_date,
    )
    db.add(review)
    db.commit()
    db.refresh(review)

    return ReviewResponse(
        id=review.id,
        card_id=card_id,
        rating=body.rating,
        interval_days=new_interval,
        ease_factor=new_ease,
        due_date=due_date.isoformat() + "Z",
    ).model_dump()


@router.get("/study/due")
def get_due_cards(
    request: Request,
    folder_id: Optional[str] = None,
    limit: int = 50,
    db: Session = Depends(get_db),
):
    """오늘 복습할 카드 목록 (due_date <= now + 미복습 카드)."""
    from datetime import datetime as dt
    from sqlalchemy import func

    owner_filter = get_owner_filter(request)
    now = dt.utcnow()

    # accepted 카드만 대상
    base_query = owner_filter(
        db.query(CardModel)
        .join(SessionModel, CardModel.session_id == SessionModel.id)
        .filter(CardModel.status == "accepted")
    )

    if folder_id:
        base_query = base_query.filter(SessionModel.folder_id == folder_id)

    all_cards = base_query.all()

    due_cards = []
    new_cards = []

    for card in all_cards:
        latest_review = (
            db.query(CardReviewModel)
            .filter(CardReviewModel.card_id == card.id)
            .order_by(CardReviewModel.reviewed_at.desc())
            .first()
        )

        if latest_review is None:
            new_cards.append(card)
        elif latest_review.due_date <= now:
            due_cards.append((card, latest_review))

    # 복습 카드 우선, 그다음 새 카드
    result = []
    for card, review in due_cards[:limit]:
        resp = _card_to_response(card)
        resp["due_date"] = review.due_date.isoformat() + "Z"
        resp["interval_days"] = review.interval_days
        resp["ease_factor"] = review.ease_factor
        resp["session_filename"] = card.session.filename if card.session else ""
        result.append(resp)

    remaining = limit - len(result)
    for card in new_cards[:remaining]:
        resp = _card_to_response(card)
        resp["due_date"] = None
        resp["interval_days"] = 0
        resp["ease_factor"] = 2.5
        resp["session_filename"] = card.session.filename if card.session else ""
        result.append(resp)

    return result


@router.get("/study/stats")
def get_study_stats(
    request: Request,
    db: Session = Depends(get_db),
    device_id: str = Depends(get_device_id),
):
    """학습 통계: 오늘 복습 수, 마스터 카드, 스트릭, 복습 대기 카드."""
    from datetime import datetime as dt

    owner = get_owner_id(request)
    now = dt.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    # 소유자 필터
    if owner["user_id"]:
        owner_cond = CardReviewModel.user_id == owner["user_id"]
    else:
        owner_cond = CardReviewModel.device_id == device_id

    # 오늘 복습 수
    reviews_today = (
        db.query(CardReviewModel)
        .filter(owner_cond, CardReviewModel.reviewed_at >= today_start)
        .count()
    )

    # 마스터 카드 (interval >= 21일, 가장 최근 복습 기준)
    from sqlalchemy import func
    latest_reviews = (
        db.query(
            CardReviewModel.card_id,
            func.max(CardReviewModel.reviewed_at).label("latest"),
        )
        .filter(owner_cond)
        .group_by(CardReviewModel.card_id)
        .subquery()
    )
    mastered = (
        db.query(CardReviewModel)
        .join(
            latest_reviews,
            (CardReviewModel.card_id == latest_reviews.c.card_id)
            & (CardReviewModel.reviewed_at == latest_reviews.c.latest),
        )
        .filter(CardReviewModel.interval_days >= 21)
        .count()
    )

    # 스트릭 (연속 학습일)
    all_review_dates = (
        db.query(func.date(CardReviewModel.reviewed_at))
        .filter(owner_cond)
        .distinct()
        .order_by(func.date(CardReviewModel.reviewed_at).desc())
        .all()
    )

    streak = 0
    expected_date = now.date()
    for (review_date,) in all_review_dates:
        if isinstance(review_date, str):
            from datetime import date
            review_date = date.fromisoformat(review_date)
        if review_date == expected_date:
            streak += 1
            expected_date -= timedelta(days=1)
        elif review_date == expected_date + timedelta(days=1):
            # 오늘 아직 안 했으면 어제부터 카운트
            if streak == 0:
                expected_date = review_date
                streak += 1
                expected_date -= timedelta(days=1)
            else:
                break
        else:
            break

    # 복습 대기 카드 수
    owner_filter = get_owner_filter(request)
    all_accepted = (
        owner_filter(
            db.query(CardModel)
            .join(SessionModel, CardModel.session_id == SessionModel.id)
            .filter(CardModel.status == "accepted")
        )
        .all()
    )

    due_count = 0
    for card in all_accepted:
        latest = (
            db.query(CardReviewModel)
            .filter(CardReviewModel.card_id == card.id)
            .order_by(CardReviewModel.reviewed_at.desc())
            .first()
        )
        if latest is None or latest.due_date <= now:
            due_count += 1

    return StudyStatsResponse(
        reviews_today=reviews_today,
        mastered_cards=mastered,
        streak_days=streak,
        due_cards=due_count,
    ).model_dump()


# ──────────────────────────────────────
# Explore — 공개 카드셋 탐색
# ──────────────────────────────────────

CATEGORIES = {
    "language": {"name": "어학", "icon": "translate", "subcategories": ["JLPT", "TOEIC", "HSK"]},
    "it": {"name": "IT/컴퓨터", "icon": "computer", "subcategories": ["정보처리기사", "컴활", "SQLD"]},
    "law": {"name": "법/행정", "icon": "gavel", "subcategories": ["행정사", "공인중개사"]},
    "business": {"name": "경영/경제", "icon": "business", "subcategories": ["한경TESAT", "매경TEST"]},
    "education": {"name": "교육", "icon": "school", "subcategories": ["교육학", "교원임용"]},
    "etc": {"name": "기타", "icon": "category", "subcategories": []},
}


@router.get("/explore/categories")
def list_categories(db: Session = Depends(get_db)):
    """카테고리 목록 + 각 카테고리별 published 카드셋 수."""
    result = []
    for key, info in CATEGORIES.items():
        count = (
            db.query(PublicCardsetModel)
            .filter(PublicCardsetModel.category == key, PublicCardsetModel.status == "published")
            .count()
        )
        result.append({
            "id": key,
            "name": info["name"],
            "icon": info["icon"],
            "subcategories": info["subcategories"],
            "cardset_count": count,
        })
    return result


@router.get("/explore/cardsets")
def list_cardsets(
    category: Optional[str] = None,
    sort: str = "popular",
    search: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """공개 카드셋 목록 (카테고리/정렬/검색 필터)."""
    query = db.query(PublicCardsetModel).filter(PublicCardsetModel.status == "published")

    if category:
        query = query.filter(PublicCardsetModel.category == category)
    if search:
        query = query.filter(PublicCardsetModel.title.like(f"%{search}%"))

    if sort == "latest":
        query = query.order_by(PublicCardsetModel.created_at.desc())
    else:
        query = query.order_by(PublicCardsetModel.download_count.desc())

    cardsets = query.limit(50).all()
    return [
        {
            "id": cs.id,
            "title": cs.title,
            "description": cs.description,
            "category": cs.category,
            "tags": cs.tags,
            "card_count": cs.card_count,
            "download_count": cs.download_count,
            "author_name": cs.author_name,
            "is_featured": cs.is_featured,
            "created_at": cs.created_at.isoformat() + "Z",
        }
        for cs in cardsets
    ]


@router.get("/explore/cardsets/{cardset_id}")
def get_cardset(cardset_id: str, db: Session = Depends(get_db)):
    """공개 카드셋 상세 + 카드 전체."""
    cs = db.query(PublicCardsetModel).filter(PublicCardsetModel.id == cardset_id).first()
    if not cs:
        raise HTTPException(404, "카드셋을 찾을 수 없습니다.")

    cards = (
        db.query(PublicCardModel)
        .filter(PublicCardModel.cardset_id == cardset_id)
        .order_by(PublicCardModel.sort_order)
        .all()
    )
    return {
        "id": cs.id,
        "title": cs.title,
        "description": cs.description,
        "category": cs.category,
        "tags": cs.tags,
        "card_count": cs.card_count,
        "download_count": cs.download_count,
        "author_name": cs.author_name,
        "is_featured": cs.is_featured,
        "created_at": cs.created_at.isoformat() + "Z",
        "cards": [
            {
                "id": c.id,
                "front": c.front,
                "back": c.back,
                "evidence": c.evidence,
                "template_type": c.template_type,
                "sort_order": c.sort_order,
            }
            for c in cards
        ],
    }


@router.post("/explore/cardsets/{cardset_id}/download")
def download_cardset(
    cardset_id: str,
    request: Request,
    db: Session = Depends(get_db),
    device_id: str = Depends(get_device_id),
):
    """공개 카드셋을 유저의 세션+카드로 복사 (로그인 필수)."""
    owner = get_owner_id(request)
    if not owner["user_id"]:
        raise HTTPException(401, "로그인이 필요합니다.")

    cs = db.query(PublicCardsetModel).filter(PublicCardsetModel.id == cardset_id).first()
    if not cs:
        raise HTTPException(404, "카드셋을 찾을 수 없습니다.")

    pub_cards = (
        db.query(PublicCardModel)
        .filter(PublicCardModel.cardset_id == cardset_id)
        .order_by(PublicCardModel.sort_order)
        .all()
    )

    # 유저 세션 생성
    session = SessionModel(
        filename=cs.title,
        page_count=0,
        template_type=pub_cards[0].template_type if pub_cards else "definition",
        device_id=device_id,
        user_id=owner["user_id"],
        source_type="explore",
        status="completed",
    )
    db.add(session)
    db.flush()

    # 카드 복사
    for pc in pub_cards:
        card = CardModel(
            session_id=session.id,
            front=pc.front,
            back=pc.back,
            evidence=pc.evidence,
            evidence_page=0,
            template_type=pc.template_type,
            status="accepted",
        )
        db.add(card)

    # download_count 증가
    cs.download_count += 1
    db.commit()
    db.refresh(session)

    return _build_session_response(session)


@router.post("/explore/publish")
def publish_cardset(
    body: PublishRequest,
    request: Request,
    db: Session = Depends(get_db),
):
    """유저 세션의 accepted 카드를 공개 카드셋으로 발행 (로그인 필수)."""
    owner = get_owner_id(request)
    if not owner["user_id"]:
        raise HTTPException(401, "로그인이 필요합니다.")

    owner_filter = get_owner_filter(request)
    session = owner_filter(
        db.query(SessionModel).filter(SessionModel.id == body.session_id)
    ).first()
    if not session:
        raise HTTPException(404, "세션을 찾을 수 없습니다.")

    accepted_cards = [c for c in session.cards if c.status == "accepted"]
    if not accepted_cards:
        raise HTTPException(400, "채택된 카드가 없습니다.")

    # 유저 정보 조회
    from .models import UserModel
    user = db.query(UserModel).filter(UserModel.id == owner["user_id"]).first()
    author_name = user.nickname if user and user.nickname else "데카드"

    category = body.category if body.category in CATEGORIES else "etc"

    cardset = PublicCardsetModel(
        title=body.title,
        description=body.description,
        category=category,
        author_id=owner["user_id"],
        author_name=author_name,
        card_count=len(accepted_cards),
    )
    db.add(cardset)
    db.flush()

    for idx, card in enumerate(accepted_cards):
        pub_card = PublicCardModel(
            cardset_id=cardset.id,
            front=card.front,
            back=card.back,
            evidence=card.evidence,
            template_type=card.template_type,
            sort_order=idx,
        )
        db.add(pub_card)

    db.commit()
    db.refresh(cardset)

    return {
        "id": cardset.id,
        "title": cardset.title,
        "description": cardset.description,
        "category": cardset.category,
        "card_count": cardset.card_count,
        "author_name": cardset.author_name,
        "created_at": cardset.created_at.isoformat() + "Z",
    }


# ──────────────────────────────────────
# Helpers
# ──────────────────────────────────────

def _safe_error(detail: str, exc: Exception) -> str:
    """Hide internal error details in production."""
    if settings.APP_ENV == "dev":
        return f"{detail}: {exc}"
    return detail


def _card_to_response(card: CardModel) -> dict:
    return CardResponse(
        id=card.id,
        front=card.front,
        back=card.back,
        evidence=card.evidence,
        evidence_page=card.evidence_page,
        tags=card.tags,
        template_type=card.template_type,
        status=card.status,
    ).model_dump()


def _build_session_response(session: SessionModel) -> dict:
    cards = [_card_to_response(c) for c in session.cards]
    stats = {
        "total": len(cards),
        "accepted": sum(1 for c in cards if c["status"] == "accepted"),
        "rejected": sum(1 for c in cards if c["status"] == "rejected"),
        "pending": sum(1 for c in cards if c["status"] == "pending"),
    }
    resp = {
        "id": session.id,
        "filename": session.filename,
        "page_count": session.page_count,
        "template_type": session.template_type,
        "status": session.status,
        "folder_id": session.folder_id,
        "display_name": session.display_name,
        "source_type": session.source_type or "pdf",
        "progress": session.progress or 0,
        "total_chunks": session.total_chunks or 0,
        "completed_chunks": session.completed_chunks or 0,
        "card_count": len(cards),
        "created_at": session.created_at.isoformat() + "Z",
        "cards": cards,
        "stats": stats,
    }
    if session.error_message:
        resp["error_message"] = session.error_message
    return resp
